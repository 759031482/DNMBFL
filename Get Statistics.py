import datetime
import math
import os
import openpyxl
from openpyxl.styles import Font  # 导入字体模块
from openpyxl.styles import PatternFill  # 导入填充模块
import matplotlib.pyplot as plt

from result_get.boxplot import Box
from result_get.boxplot import Violin
from result_get.boxplot import Bline
from result_get.boxplot import MutiBline
# from result_get.boxplot import Line



from codeflaws_version_control import Qr_excel

suslist = [
    'Tarantula',
    'Op2',
    'Jaccard',
    'Ochiai',
    'Dstar',
    'GP13',
    'Naish1',
    'Barinel',
]

def readxlsx(path, sheet):
    try:
        data_read = Qr_excel().read(path, sheet)
    except:
        print('读取失败 %s %s' % (path, sheet))
        return False
    static_all = {
        'totalfaultnum': 0,
        'ttime': 0,
        'fomnum': 0,
        'homnum': 0,
        'Operator': [0 for _ in range(13)],
        'Accurate': [0 for _ in range(3)],
        'versionnum': len(data_read)-1,
        'form': dict()
    }
    for i in range(len(data_read)):
        row_data = data_read[i]
        for key, value in row_data['self'].items():
            if value == '':
                continue
            if key == 'Fault_Record':
                static_all['totalfaultnum'] += int(value)
            elif key == 'TimeSpend':
                static_all['ttime'] += int(value)
            elif key == 'Fomnum':
                try:
                    static_all['fomnum'] += int(value)
                except:
                    print(row_data['self'])
            elif key == 'Homnum':
                static_all['homnum'] += int(value)
            elif 'Operator' in key:
                OperatorType = int(key.split(' ')[2])
                static_all['Operator'][OperatorType] = static_all['Operator'][OperatorType] + value*100
            elif 'Accurate' in key:
                if key == 'None Accurate':
                    static_all['Accurate'][0] += value*100
                if key == 'Pare Accurate':
                    static_all['Accurate'][1] += value*100
                if key == 'Accurate':
                    static_all['Accurate'][2] += value*100
        for key, value in row_data['form'].items():
            if key not in static_all['form']:
                static_all['form'][key] = {
                    'exam': 0,
                    # 'exams': dict(),
                    'exams': [],
                    'rank': [0 for _ in range(4)],
                    'map': 0,
                }
            for j in range(len(row_data['form'][key]['exam'])):
                exam = row_data['form'][key]['exam'][j]
                rank = row_data['form'][key]['rank'][j]
                static_all['form'][key]['exam'] += exam
                # if j == 0:
                #     static_all['form'][key]['exams'][row_data['self']['Version']] = [exam]
                # else:
                #     static_all['form'][key]['exams'][row_data['self']['Version']].append(exam)
                static_all['form'][key]['exams'].append(exam)
                # static_all['form'][key]['exams'].append(exam)
                static_all['form'][key]['map'] += 1/rank
                if rank <= 1:
                    static_all['form'][key]['rank'][0] += 1
                if rank <= 3:
                    static_all['form'][key]['rank'][1] += 1
                if rank <= 5:
                    static_all['form'][key]['rank'][2] += 1
                if rank <= 10:
                    static_all['form'][key]['rank'][3] += 1

    static_all['fomnum'] = static_all['fomnum']/static_all['versionnum']
    static_all['homnum'] = static_all['homnum']/static_all['versionnum']
    for i in range(13):
        static_all['Operator'][i] = static_all['Operator'][i]/static_all['versionnum']
    for i in range(3):
        static_all['Accurate'][i] = static_all['Accurate'][i]/static_all['versionnum']
    for fun in static_all['form'].keys():
        if static_all['totalfaultnum'] == 0:
            static_all['form'][fun]['exam'] = 0
            static_all['form'][fun]['map'] = 0
        else:
            static_all['form'][fun]['exam'] = static_all['form'][fun]['exam']/static_all['totalfaultnum']
            static_all['form'][fun]['map'] = static_all['form'][fun]['map']/static_all['versionnum']

    return static_all

class Top:
    def total(self):
        color = ['006100', '9c0006', '9c6500', '000000']
        # 绿 红 黄 黑
        savepath = os.path.join(os.getcwd(), r'report/CHMBFL/susfile/static',
                                'top DeltaNs-4.xlsx')
        baseline_list = [
            ['Random1', 'numless/Random/1'],
            ['Random2', 'numless/Random/2'],
            ['Random3', 'numless/Random/3'],
            ['Random4', 'numless/Random/4'],
            ['Random5', 'numless/Random/5'],
            ['Random6', 'numless/Random/6'],
            ['Random7', 'numless/Random/7'],
            ['Random8', 'numless/Random/8'],
            ['Random9', 'numless/Random/9'],
            ['Random10', 'numless/Random/10'],
            ['Random11', 'numless/Random/11'],
            ['Random12', 'numless/Random/12'],
            ['Random13', 'numless/Random/13'],
            ['Random14', 'numless/Random/14'],
            ['Random15', 'numless/Random/15'],
            ['Random16', 'numless/Random/16'],
            ['Random17', 'numless/Random/17'],
            ['Random18', 'numless/Random/18'],
            ['Random19', 'numless/Random/19'],
        ]
        baseline_list = [
            # ['NS-MS (10', '', 'numless/NS-MutCluster-DifClass (10', color[3]],
            # ['NS (10', '', 'numless/NS (10', color[3]],

            # ['Sbfl', 'baselines', 'baseline/Sbfl', color[3]],
            # ['Mbfl', 'baselines', 'baseline/Mbfl', color[3]],
            # ['Mcbfl', 'baselines', 'baseline/Mcbfl', color[3]],
            # ['Last2First', 'baselines', 'baseline/Last2First', color[3]],
            # ['DifferentOperator', 'baselines', 'baseline/DifferentOperator', color[3]],
            # ['RandomMix', 'baselines', 'baseline/RandomMix', color[3]],
            # ['Random', 'baselines', 'baseline/Random', color[3]],

            # ['DeltaMbfl', '', 'baseline/DeltaMbfl', color[1]],
            # ['DeltaNS', '使用Delta4Ms的非叠加变异', 'numless/DeltaNS', color[2]],
            ['DeltaNsMbfl', '', 'numless/DeltaNsMbfl-1', color[2]],
            # ['ED', '基于错误分布的高阶变异体生成方法', 'numless/ErrorDistribution', color[1]],
            # ['ED.SBFL', 'SBFL指导的基于错误分布的高阶变异体生成方法', 'numless/ED.SBFL', color[1]],
            # ['FTC', '基于失败测试用例聚类的高阶变异体生成方法(mseer)', 'numless/FailTestCluster', color[1]],
            # ['FTC.kmeans', '基于失败测试用例聚类的高阶变异体生成方法(kmeans)', 'numless/FailTestCluster.kmeans', color[1]],
            # ['MC', '基于变异体聚类的高阶变异体生成方法(mseer)', 'numless/MutCluster-DifClass', color[1]],
            # ['MC.in', '基于变异体聚类的高阶变异体生成方法(mseer)(类簇内）', 'numless/MutCluster.in', color[1]],
            # ['MC.kmeans', '基于变异体聚类的高阶变异体生成方法(kmeans)', 'numless/MutCluster.kmeans', color[1]],
            # ['MC.kmeans.in', '基于变异体聚类的高阶变异体生成方法(kmeans)(类簇内）', 'numless/MutCluster.kmeans.in', color[1]],
            # ['NS（随机）', '非叠加变异', 'numless/NS (10', color[1]],
            # ['NS.SBFL', 'SBFL指导的非叠加变异', 'numless/NS.SBFL', color[1]],
            # ['NS.MBFL', 'MBFL指导的非叠加变异', 'numless/NS.MBFL', color[1]],
            # ['MC', '基于变异体聚类的高阶变异体生成方法', 'numless/MutCluster-DifClass', color[1]],
            # ['RS', '基于变异体怀疑度重组的高阶变异体生成方法', 'numless/RegroupBySus-H', color[1]]
            # ['NS.mbfl%FTC(mseer)', '组合方法 非叠加变异-基于失败测试用例聚类(各取50%)', 'numless/NS.mbfl%25FTC', color[0]],
            # ['NS.mbfl%MC(mseer)', '组合方法 非叠加变异-基于变异体聚类(各取50%)', 'numless/NS.mbfl%25MC', color[0]],
            # ['NS.mbfl%ErrorDistribution', '组合方法 非叠加变异-基于错误分布(各取50%)', 'numless/NS.mbfl%25ED', color[0]],
            # ['MC(mseer)%ED', '组合方法 基于变异体聚类-基于错误分布 集合各取50%', 'numless/MC%25ED', color[0]],
            # ['ED%NS%MC(mseer)', '组合方法  集合各取1/3', 'numless/ED%25NS%25MC.mseer', color[0]],
            # ['ED%NS%FTC(mseer)', '组合方法  集合各取1/3', 'numless/ED%25NS%25FTC.mseer', color[0]],
            # ['NS-FTC', '组合方法 非叠加变异-基于失败测试用例聚类', 'numless/NS-FailTestCluster', color[0]],
            # ['NS-MC', '组合方法 非叠加变异-基于变异体聚类', 'numless/NS-MutCluster-DifClass', color[0]],
            # ['NS-RegroupBySus', '组合方法 非叠加变异-基于变异体怀疑度重组', 'numless/NS-RegroupBySus-H', color[0]],
            # ['NS-ErrorDistribution', '组合方法 非叠加变异-基于错误分布', 'numless/NS-ErrorDistribution', color[0]],
            # ['MC+FTC', '组合方法 基于变异体聚类-基于失败测试用例聚类 集合取交集', 'numless/MC+FTC', color[0]],
            # ['MC%FTC', '组合方法 基于变异体聚类-基于失败测试用例聚类 集合各取50%', 'numless/MC%FTC', color[0]],
            # ['MC+ED', '组合方法 基于变异体聚类-基于错误分布 集合取交集', 'numless/MC+ED', color[0]],
            # ['MC%ED', '组合方法 基于变异体聚类-基于错误分布 集合各取50%', 'numless/MC%ED', color[0]],
            # ['NS-RS-ED', '组合方法 非叠加变异-基于变异体怀疑度重组-基于错误分布', 'numless/NS-RS-ED', color[2]],
            # ['NS-RS-FTC', '组合方法 非叠加变异-基于变异体怀疑度重组-基于失败测试用例聚类', 'numless/NS-RS-FTC', color[2]],
            # ['NS-RS-MC', '组合方法 非叠加变异-基于变异体怀疑度重组-基于变异体聚类', 'numless/NS-RS-MC', color[2]],
            # ['RS-MC+ED', '组合方法 基于变异体怀疑度重组-（基于错误分布 与 基于错误分布组合）', 'numless/RS-MC+ED', color[2]],
            # ['RS-MC+FTC', '组合方法 基于变异体怀疑度重组-（基于错误分布 与 基于失败测试用例聚类）', 'numless/RS-MC+FTC', color[2]],
            # ['NS-RS-MC+ED', '组合方法 非叠加变异 基于变异体怀疑度重组-（基于变异体聚类 与 基于错误分布）', 'numless/NS-RS-MC+ED', color[2]],
            # ['NS-RS-MC+FTC', '组合方法 非叠加变异 基于变异体怀疑度重组-（基于变异体聚类 与 基于失败测试用例聚类）', 'numless/NS-RS-MC+FTC', color[2]],
            # ['TestAnalysis', '基于测试评估的高阶变异体约简方法', 'numless/TestAnalysis', color[2]],
        ]
        endnum = 0
        tnum = len(baseline_list)*3
        sustype_list = ['max', 'ave', 'frequency']
        tietype_list = ['exam-best', 'exam-average', 'exam-worst']

        wb = openpyxl.Workbook()
        del wb['Sheet']
        # ws = wb['Sheet']
        Wss = [wb.create_sheet('best'), wb.create_sheet('ave'), wb.create_sheet('worst'), ]
        line = 1
        sheettitle = ['baseline',
                      'describe',
                      # 'classnum',
                      'sustype',
                      'tietype',
                      'method',
                      'totalfaultnum',
                      'ttime',
                      'versionnum',
                      'K',
                      'fomnum',
                      'homnum',
                      'map',
                      'exam',
                      'top1',
                      'top3',
                      'top5',
                      'top10',
                      '>top10']
        for ws in Wss:
            for i, title in enumerate(sheettitle):
                ws.cell(line, i+1, title)

        stime = datetime.datetime.now()
        for tie_i, tietype in enumerate(tietype_list):
            line = 2
            ws = Wss[tie_i]
            for baseline, describe, baseline_path, color in baseline_list:
                endnum += 1
                nowtime = datetime.datetime.now()
                if endnum == 1:
                    print("%s %.2f%s, 剩余时间:%s" % (nowtime, endnum/tnum*100, "%", 'unknown'))
                else:
                    print("%s %.2f%s, 剩余时间:%s, 结束时间:%s" %
                          (nowtime, endnum/tnum*100, "%",
                           (nowtime-stime)/(endnum-1)*(tnum-endnum+1),
                           stime+(nowtime-stime)/(endnum-1)*tnum,))
                for sustype in sustype_list:
                    path = os.path.join(os.getcwd(), r'report/CHMBFL/susfile', baseline_path, sustype)
                    if not sustype == 'max' and (baseline == 'Mbfl' or baseline == 'Mcbfl' or baseline == 'DeltaMbfl'):
                        path = os.path.join(os.getcwd(), r'report/CHMBFL/susfile', baseline_path, 'max')
                    if baseline == 'Sbfl':
                        path = os.path.join(os.getcwd(), r'report/CHMBFL/susfile', baseline_path)

                    if not os.path.exists(path):
                        print('路径不存在 %s' % path)
                        continue
                    staticdic = dict()

                    static_list = []
                    for file in os.listdir(path):
                        filepath = os.path.join(path, file)
                        static = readxlsx(filepath, tietype)
                        if not static:
                            continue
                        static_list.append(static)

                    for static in static_list:
                        for method, value in static.items():
                            if method not in staticdic:
                                staticdic[method] = {
                                    'ranks': [0, 0, 0, 0, 0],
                                    'totalfaultnum': 0,
                                    'ttime': 0,
                                    'versionnum': 0,
                                    'K': 0,
                                    'fomnum': 0,
                                    'homnum': 0,
                                    'exams': [],
                                    'map': [],
                                }
                            staticdic[method]['ranks'][0] += value['ranks'][0]
                            staticdic[method]['ranks'][1] += value['ranks'][1]
                            staticdic[method]['ranks'][2] += value['ranks'][2]
                            staticdic[method]['ranks'][3] += value['ranks'][3]
                            staticdic[method]['ranks'][4] += value['ranks'][4]
                            staticdic[method]['totalfaultnum'] += value['totalfaultnum']
                            staticdic[method]['ttime'] += value['ttime']
                            staticdic[method]['versionnum'] += value['versionnum']
                            staticdic[method]['K'] += value['K']
                            staticdic[method]['fomnum'] += value['fomnum']
                            staticdic[method]['homnum'] += value['homnum']
                            staticdic[method]['exams'] += value['exams']
                            staticdic[method]['map'] += value['map']

                    for method in staticdic.keys():
                        for i in range(5):
                            staticdic[method]['ranks'][i] = math.ceil(staticdic[method]['ranks'][i]/len(static_list))
                        staticdic[method]['totalfaultnum'] = staticdic[method]['totalfaultnum'] /len(static_list)
                        staticdic[method]['ttime'] = staticdic[method]['ttime']/len(static_list)
                        staticdic[method]['versionnum'] = staticdic[method]['versionnum']/len(static_list)
                        staticdic[method]['K'] = staticdic[method]['K']/len(static_list)
                        staticdic[method]['exams'] = sum(staticdic[method]['exams'])/len(staticdic[method]['exams'])
                        staticdic[method]['map'] = sum(staticdic[method]['map'])/len(staticdic[method]['map'])

                    for method in staticdic.keys():
                        staticdata = [
                            baseline,
                            describe,
                            sustype,
                            tietype,
                            method,
                            staticdic[method]['totalfaultnum'],
                            staticdic[method]['ttime'],
                            staticdic[method]['versionnum'],
                            float('%.2f' % staticdic[method]['K']),
                            float('%.2f' % staticdic[method]['fomnum']),
                            float('%.2f' % staticdic[method]['homnum']),
                            float('%.4f' % staticdic[method]['map']),
                            float('%.4f' % staticdic[method]['exams']),
                            staticdic[method]['ranks'][0],
                            staticdic[method]['ranks'][1],
                            staticdic[method]['ranks'][2],
                            staticdic[method]['ranks'][3],
                            staticdic[method]['ranks'][4],
                        ]
                        font = Font(u'微软雅黑', size=11, bold=True, italic=False, strike=False, color=color)
                        for i, data in enumerate(staticdata):
                            # ws.cell(row=line, column=i+1, value=data)
                            ws.cell(row=line, column=i+1, value=data).font = font
                            # ws.cell(line, i+1, data).font = font
                        line += 1

        wb.save(savepath)
        print('save %s' % savepath)
        return

    def total_type(self):
        color = ['006100', '9c0006', '9c6500', '000000']
        # 绿 红 黄 黑
        doc_path = './report/CHMBFL/susfile-single'
        savepath = os.path.join(os.getcwd(), doc_path, 'static',
                                'codeflaws_single.xlsx')
        baseline_list = [

            ['Sbfl', 'baselines', 'baseline/Sbfl', color[3]],
            ['Mbfl', 'baselines', 'baseline/Mbfl', color[3]],
            ['Mcbfl', 'baselines', 'baseline/Mcbfl', color[3]],
            ['Last2First', 'baselines', 'HomBaseline/Last2First-0_5', color[3]],
            ['DifferentOperator', 'baselines', 'HomBaseline/DifferentOperator-0_5', color[3]],
            ['RandomMix', 'baselines', 'HomBaseline/RandomMix-0_5', color[3]],
            ['Random', 'baselines', 'HomBaseline/Random-0_5', color[3]],

            ['DeltaNsMbfl-0', '', 'DNMbfl/DeltaNsMbfl-0', color[2]],
            ['DeltaNsMbfl-0.5', '', 'DNMbfl/DeltaNsMbfl-0.5', color[2]],
            ['DeltaNsMbfl-1', '', 'DNMbfl/DeltaNsMbfl-1', color[2]],
            ['DeltaNsMbfl-1.5', '', 'DNMbfl/DeltaNsMbfl-1.5', color[2]],
            ['DeltaNsMbfl-2', '', 'DNMbfl/DeltaNsMbfl-2', color[2]],
            ['DeltaNsMbfl-2.5', '', 'DNMbfl/DeltaNsMbfl-2.5', color[2]],
            ['DeltaNsMbfl-3', '', 'DNMbfl/DeltaNsMbfl-3', color[2]],
            ['DeltaNsMbfl-3.5', '', 'DNMbfl/DeltaNsMbfl-3.5', color[2]],
            ['DeltaNsMbfl-4', '', 'DNMbfl/DeltaNsMbfl-4', color[2]],

        ]
        # baseline_list = [
        #     ['DNMBFL-0', '', 'DNMbfl/DeltaNsMbfl-0', color[2]],
        #     ['DNMBFL-0.5', '', 'DNMbfl/DeltaNsMbfl-0.5', color[2]],
        #     ['DNMBFL-1', '', 'DNMbfl/DeltaNsMbfl-1', color[2]],
        #     ['DNMBFL-1.5', '', 'DNMbfl/DeltaNsMbfl-1.5', color[2]],
        #     ['DNMBFL-2', '', 'DNMbfl/DeltaNsMbfl-2', color[2]],
        #     ['DNMBFL-2.5', '', 'DNMbfl/DeltaNsMbfl-2.5', color[2]],
        #     ['DNMBFL-3', '', 'DNMbfl/DeltaNsMbfl-3', color[2]],
        #     ['DNMBFL-3.5', '', 'DNMbfl/DeltaNsMbfl-3.5', color[2]],
        #     ['DNMBFL-4', '', 'DNMbfl/DeltaNsMbfl-4', color[2]]
        # ]
        endnum = 0
        tnum = len(baseline_list)*3
        sustype_list = ['max', 'ave', 'frequency']
        tietype_list = ['exam-best', 'exam-average', 'exam-worst']

        wb = openpyxl.Workbook()
        del wb['Sheet']
        # ws = wb['Sheet']
        Wss = [wb.create_sheet('best'), wb.create_sheet('ave'), wb.create_sheet('worst'), ]
        line = 1
        sheettitle = [
            'baseline',
            'describe',
            'sustype',
            'tietype',
            'method',
            'totalfaultnum',
            'ttime',
            'versionnum',
            'K',
            'fomnum',
            'homnum',
        ]
        for i in range(13):
            sheettitle.append('Operator Type %s' % str(i))
        for AccurateName in ['None Accurate', 'Pare Accurate', 'Accurate']:
            sheettitle.append(AccurateName)
        sheettitle += [
            'map',
            'exam',
            'top1',
            'top3',
            'top5',
            'top10',
        ]
        for ws in Wss:
            for i, title in enumerate(sheettitle):
                ws.cell(line, i+1, title)

        stime = datetime.datetime.now()
        for tie_i, tietype in enumerate(tietype_list):
            line = 2
            ws = Wss[tie_i]
            for baseline, describe, baseline_path, color in baseline_list:
                endnum += 1
                nowtime = datetime.datetime.now()
                if endnum == 1:
                    print("%s %.2f%s, 剩余时间:%s" % (nowtime, endnum/tnum*100, "%", 'unknown'))
                else:
                    print("%s %.2f%s, 剩余时间:%s, 结束时间:%s" %
                          (nowtime, endnum/tnum*100, "%",
                           (nowtime-stime)/(endnum-1)*(tnum-endnum+1),
                           stime+(nowtime-stime)/(endnum-1)*tnum,))
                for sustype in sustype_list:
                    path = os.path.join(os.getcwd(), doc_path, baseline_path, sustype)
                    if not sustype == 'max' and (baseline == 'Mbfl' or baseline == 'Mcbfl' or baseline == 'DeltaMbfl'):
                        path = os.path.join(os.getcwd(), doc_path, baseline_path, 'max')
                    if baseline == 'Sbfl':
                        path = os.path.join(os.getcwd(), doc_path, baseline_path)

                    if not os.path.exists(path):
                        print('路径不存在 %s' % path)
                        continue
                    staticdic = {
                        'totalfaultnum': 0,
                        'ttime': 0,
                        'fomnum': 0,
                        'homnum': 0,
                        'Operator': [0 for _ in range(13)],
                        'Accurate': [0 for _ in range(3)],
                        'versionnum': 0,
                        'form': dict()
                    }

                    static_list = []
                    for file in os.listdir(path):
                        filepath = os.path.join(path, file)
                        static = readxlsx(filepath, tietype)
                        if not static:
                            continue
                        static_list.append(static)

                    for static in static_list:
                        for key, value in static.items():
                            if key == 'totalfaultnum':
                                staticdic['totalfaultnum'] += value
                            if key == 'ttime':
                                staticdic['ttime'] += value
                            if key == 'fomnum':
                                staticdic['fomnum'] += value
                            if key == 'homnum':
                                staticdic['homnum'] += value
                            if key == 'versionnum':
                                staticdic['versionnum'] += value
                            if key == 'Operator':
                                for i in range(13):
                                    staticdic['Operator'][i] += value[i]
                            if key == 'Accurate':
                                for i in range(3):
                                    # try:
                                    staticdic['Accurate'][i] += value[i]
                                    # except:
                                    #     print('1')
                            if key == 'form':
                                for form in value.keys():
                                    if form not in staticdic['form']:
                                        staticdic['form'][form] = {
                                            'exam': 0,
                                            'rank': [0 for _ in range(4)],
                                            'map': 0,
                                        }
                                    staticdic['form'][form]['exam'] += value[form]['exam']
                                    staticdic['form'][form]['map'] += value[form]['map']
                                    for i in range(4):
                                        staticdic['form'][form]['rank'][i] += value[form]['rank'][i]

                    for method in staticdic['form'].keys():
                        staticdata = [
                            baseline,
                            describe,
                            sustype,
                            tietype,
                            method,
                            staticdic['totalfaultnum']/len(static_list),
                            staticdic['ttime']/len(static_list),
                            staticdic['versionnum']/len(static_list),
                            staticdic['homnum']/staticdic['fomnum'] if staticdic['fomnum'] > 0 else 0,
                            staticdic['fomnum']/len(static_list),
                            staticdic['homnum']/len(static_list),
                        ]
                        for i in range(13):
                            staticdata.append(staticdic['Operator'][i]/len(static_list))
                        for i in range(3):
                            staticdata.append(staticdic['Accurate'][i]/len(static_list))
                        staticdata.append(staticdic['form'][method]['map']/len(static_list))
                        staticdata.append(staticdic['form'][method]['exam']/len(static_list))
                        for i in range(4):
                            staticdata.append(staticdic['form'][method]['rank'][i]/len(static_list))



                        font = Font(u'微软雅黑', size=11, bold=True, italic=False, strike=False, color=color)
                        for i, data in enumerate(staticdata):
                            # ws.cell(row=line, column=i+1, value=data)
                            ws.cell(row=line, column=i+1, value=data).font = font
                            # ws.cell(line, i+1, data).font = font
                        line += 1

        wb.save(savepath)
        print('save %s' % savepath)
        return

    def total_K1(self):
        color = ['006100', '9c0006', '9c6500', '000000']
        # 绿 红 黄 黑

        baseline_list = [
            ['SBFL', 'baseline', '', color[3]],
            ['MBFL', 'baseline', '', color[3]],
            ['MCBFL', 'baseline', '', color[3]],

            ['DifferentOperator', '', 'DifferentOperator', color[2]],
            ['Last2First', '', 'Last2First', color[2]],
            ['RandomMix', '', 'RandomMix', color[2]],
            ['Random', '', 'Random', color[2]],

            ['ED.MBFL', '', 'ED.MBFL', color[1]],
            ['ED.SBFL', '', 'ED.SBFL', color[1]],
            ['FTC', '', 'FTC', color[1]],
            ['FTC.kmeans', '', 'FTC.kmeans', color[1]],
            ['MutCluster', '', 'MutCluster', color[1]],
            ['MutCluster.in', '', 'MutCluster.in', color[1]],
            ['MutCluster.kmeans', '', 'MutCluster.kmeans', color[1]],
            ['MutCluster.kmeans.in', '', 'MutCluster.kmeans.in', color[1]],
            ['NS.RANDOM', '', 'NS.RANDOM', color[1]],
            ['NS.SBFL', '', 'NS.SBFL', color[1]],
            ['NS.MBFL', '', 'NS.MBFL', color[1]],

            ['NS.mbfl^FTC', '', 'NS.mbfl^FTC', color[0]],
            ['NS.mbfl^MC', '', 'NS.mbfl^MC', color[0]],
            ['NS.mbfl^ED', '', 'NS.mbfl^ED', color[0]],
            ['NS.mbfl^ED', '', 'NS.mbfl^ED', color[0]],
            ['MC.mseer^ED', '', 'MC.mseer^ED', color[0]],
            ['ED^NS^MC.mseer', '', 'ED^NS^MC.mseer', color[0]],
            ['ED^NS^FTC.mseer', '', 'ED^NS^FTC.mseer', color[0]],

        ]

        tnum = len(baseline_list)
        # sustype_list = ['max', 'ave', 'frequency', 'none']
        sustype_list = ['max', 'ave', 'frequency']
        tietype_list = ['exam-best', 'exam-average', 'exam-worst']

        for K in os.listdir('./report/CHMBFL/susfile/nK'):
            endnum = 0
            savepath = os.path.join(os.getcwd(), r'report/CHMBFL/susfile/static',
                                    'whf3-%s.xlsx' % K)
            wb = openpyxl.Workbook()
            del wb['Sheet']
            # ws = wb['Sheet']
            Wss = [wb.create_sheet('best'), wb.create_sheet('ave'), wb.create_sheet('worst'), ]
            line = 1
            sheettitle = [
                'baseline',
                'describe',
                'sustype',
                'tietype',
                'method',
                'totalfaultnum',
                'ttime',
                'versionnum',
                'K',
                'fomnum',
                'homnum',
            ]
            for i in range(13):
                sheettitle.append('Operator Type %s' % str(i))
            for AccurateName in ['None Accurate', 'Pare Accurate', 'Accurate']:
                sheettitle.append(AccurateName)
            sheettitle += [
                'map',
                'exam',
                'top1',
                'top3',
                'top5',
                'top10',
            ]
            for ws in Wss:
                for i, title in enumerate(sheettitle):
                    ws.cell(line, i+1, title)

            stime = datetime.datetime.now()
            for tie_i, tietype in enumerate(tietype_list):
                line = 2
                ws = Wss[tie_i]
                for baseline, describe, baseline_path, color in baseline_list:
                    endnum += 1
                    nowtime = datetime.datetime.now()
                    if endnum == 1:
                        print(K, "  %s %.2f%s, 剩余时间:%s" % (nowtime, endnum/tnum*100, "%", 'unknown'))
                    else:
                        print(K, "  %s %.2f%s, 剩余时间:%s, 结束时间:%s" %
                              (nowtime, endnum/tnum*100, "%",
                               (nowtime-stime)/(endnum-1)*(tnum-endnum+1),
                               stime+(nowtime-stime)/(endnum-1)*tnum,))
                    for sustype in sustype_list:
                        if baseline == 'SBFL':
                            path = os.path.join(os.getcwd(), r'report/CHMBFL/susfile/baseline/Sbfl')
                        elif baseline == 'MBFL':
                            path = os.path.join(os.getcwd(), r'report/CHMBFL/susfile/baseline/Mbfl/max')
                        elif baseline == 'MCBFL':
                            path = os.path.join(os.getcwd(), r'report/CHMBFL/susfile/baseline/Mcbfl/max')
                        elif baseline in ['DifferentOperator', 'Last2First', 'RandomMix', 'Random']:
                            path = os.path.join(os.getcwd(), r'report/CHMBFL/susfile/HomBaseline', '%s-%s'% (baseline_path, '0_5'), sustype)
                        else:
                            path = os.path.join(os.getcwd(), r'report/CHMBFL/susfile/nK', K, '%s-%s'% (baseline_path, K), sustype)

                        if not os.path.exists(path):
                            print('路径不存在 %s' % path)
                            continue
                        staticdic = {
                            'totalfaultnum': 0,
                            'ttime': 0,
                            'fomnum': 0,
                            'homnum': 0,
                            'Operator': [0 for _ in range(13)],
                            'Accurate': [0 for _ in range(3)],
                            'versionnum': 0,
                            'form': dict()
                        }

                        static_list = []
                        for file in os.listdir(path):
                            filepath = os.path.join(path, file)
                            static = readxlsx(filepath, tietype)
                            if not static:
                                continue
                            static_list.append(static)

                        for static in static_list:
                            for key, value in static.items():
                                if key == 'totalfaultnum':
                                    staticdic['totalfaultnum'] += value
                                if key == 'ttime':
                                    staticdic['ttime'] += value
                                if key == 'fomnum':
                                    staticdic['fomnum'] += value
                                if key == 'homnum':
                                    staticdic['homnum'] += value
                                if key == 'versionnum':
                                    staticdic['versionnum'] += value
                                if key == 'Operator':
                                    for i in range(13):
                                        staticdic['Operator'][i] += value[i]
                                if key == 'Accurate':
                                    for i in range(3):
                                        # try:
                                        staticdic['Accurate'][i] += value[i]
                                        # except:
                                        #     print('1')
                                if key == 'form':
                                    for form in value.keys():
                                        if form not in staticdic['form']:
                                            staticdic['form'][form] = {
                                                'exam': 0,
                                                'rank': [0 for _ in range(4)],
                                                'map': 0,
                                            }
                                        staticdic['form'][form]['exam'] += value[form]['exam']
                                        staticdic['form'][form]['map'] += value[form]['map']
                                        for i in range(4):
                                            staticdic['form'][form]['rank'][i] += value[form]['rank'][i]

                        for method in staticdic['form'].keys():
                            staticdata = [
                                baseline,
                                describe,
                                sustype,
                                tietype,
                                method,
                                staticdic['totalfaultnum']/len(static_list),
                                staticdic['ttime']/len(static_list),
                                staticdic['versionnum']/len(static_list),
                                staticdic['homnum']/staticdic['fomnum'] if staticdic['fomnum'] > 0 else 0,
                                staticdic['fomnum']/len(static_list),
                                staticdic['homnum']/len(static_list),
                            ]
                            for i in range(13):
                                staticdata.append(staticdic['Operator'][i]/len(static_list))
                            for i in range(3):
                                staticdata.append(staticdic['Accurate'][i]/len(static_list))
                            staticdata.append(staticdic['form'][method]['map']/len(static_list))
                            staticdata.append(staticdic['form'][method]['exam']/len(static_list))
                            for i in range(4):
                                staticdata.append(math.ceil(staticdic['form'][method]['rank'][i]/len(static_list)))

                            font = Font(u'微软雅黑', size=11, bold=True, italic=False, strike=False, color=color)
                            for i, data in enumerate(staticdata):
                                # ws.cell(row=line, column=i+1, value=data)
                                ws.cell(row=line, column=i+1, value=data).font = font
                                # ws.cell(line, i+1, data).font = font
                            line += 1

            wb.save(savepath)
            print('save %s' % savepath)
        return

    def numneed(self):
        baseline_list = ['AgglomerativeClustering', 'KMeans']
        sustype_list = ['max', 'ave', 'frequency']
        tietype_list = ['exam-best', 'exam-average', 'exam-worst']

        wb = openpyxl.Workbook()
        ws = wb['Sheet']
        line = 1
        sheettitle = ['baseline',
                      'classnum',
                      'sustype',
                      'tietype',
                      'method',
                      'totalfaultnum',
                      'ttime',
                      'versionnum',
                      'top1',
                      'top3',
                      'top3',
                      'top10',
                      '>top10']
        for i, title in enumerate(sheettitle):
            ws.cell(line, i+1, title)
        line += 1

        for baseline in baseline_list:
            for sustype in sustype_list:
                path = os.path.join(os.getcwd(), r'report/CHMBFL/susfile/cluster_sameline', baseline, sustype)
                for doc_num in os.listdir(path):
                    l = doc_num.split('_')
                    num = doc_num.split('_')[0]
                    for tietype in tietype_list:
                        staticdic = dict()

                        static_list = []
                        for file in os.listdir(os.path.join(path, doc_num)):
                            filepath = os.path.join(path, doc_num, file)
                            static = readxlsx(filepath, tietype)
                            if not static:
                                static_list.append(static)

                        for static in static_list:
                            for method, value in static.items():
                                if method not in staticdic:
                                    staticdic[method] = {
                                        'ranks': [0, 0, 0, 0, 0],
                                        'totalfaultnum': 0,
                                        'ttime': 0,
                                        'versionnum': 0
                                    }
                                staticdic[method]['ranks'][0] += value['ranks'][0]
                                staticdic[method]['ranks'][1] += value['ranks'][1]
                                staticdic[method]['ranks'][2] += value['ranks'][2]
                                staticdic[method]['ranks'][3] += value['ranks'][3]
                                staticdic[method]['ranks'][4] += value['ranks'][4]
                                staticdic[method]['totalfaultnum'] += value['totalfaultnum']
                                staticdic[method]['ttime'] += value['ttime']
                                staticdic[method]['versionnum'] += value['versionnum']

                        for method in staticdic.keys():
                            staticdic[method]['ranks'][0] = staticdic[method]['ranks'][0]/len(static_list)
                            staticdic[method]['ranks'][1] = staticdic[method]['ranks'][1]/len(static_list)
                            staticdic[method]['ranks'][2] = staticdic[method]['ranks'][2]/len(static_list)
                            staticdic[method]['ranks'][3] = staticdic[method]['ranks'][3]/len(static_list)
                            staticdic[method]['ranks'][4] = staticdic[method]['ranks'][4]/len(static_list)
                            staticdic[method]['totalfaultnum'] = staticdic[method]['totalfaultnum']/len(static_list)
                            staticdic[method]['ttime'] = staticdic[method]['ttime']/len(static_list)
                            staticdic[method]['versionnum'] = staticdic[method]['versionnum']/len(static_list)

                        for method in staticdic.keys():
                            staticdata = [
                                baseline,
                                num,
                                sustype,
                                tietype,
                                method,
                                staticdic[method]['totalfaultnum'],
                                staticdic[method]['ttime'],
                                staticdic[method]['versionnum'],
                                staticdic[method]['ranks'][0],
                                staticdic[method]['ranks'][1],
                                staticdic[method]['ranks'][2],
                                staticdic[method]['ranks'][3],
                                staticdic[method]['ranks'][4],
                            ]
                            for i, data in enumerate(staticdata):
                                ws.cell(line, i+1, data)
                            line += 1

        savepath = os.path.join(os.getcwd(), r'report/CHMBFL/susfile/static', 'static-numneed.xlsx')
        wb.save(savepath)
        print('save %s' % savepath)
        return

class Mtp:
    def re3(self):
        data = {
            'Codeflaws Muti Faults': {
                # 'SBFl': 8463,
                'Metallaxis': 1097158,
                # 'MCBFL': 1097158,
                'MCBFL-\nhybrid-avg': 1097158,
                'Last2First': 546643,
                'Different\nOperator': 546643,
                'RandomMix': 546643,
                'Random': 546643,
                'Delta4Ms': 2205349,
            },
            'Codeflaws Single Fault': {
                # 'SBFl': 9063,
                'Metallaxis': 1257377,
                'MCBFL-\nhybrid-avg': 1257377,
                'Last2First': 626788,
                'Different\nOperator': 626788,
                'RandomMix': 626788,
                'Random': 626788,
                'Delta4Ms': 2526776,
            },
        }
        font = {'family': 'Times New Roman'}
        for cow, set in enumerate(data.keys()):
            # plt.title(set)
            # plt.subplot(len(data.keys()), 1, cow+1)
            plt.ylabel('MTP(million)', fontdict=font)
            xd = list(map(lambda s: data[set][s]/100000, data[set].keys()))
            x = list(map(lambda s: data[set][s], data[set].keys()))
            y = list(data[set].keys())
            hatch = ['-', '++++', 'xxxx', '\\\\\\', '////', '...', '.']
            color = ['white', 'white', 'white', 'white', 'white', 'white', 'black']
            plt.bar([i+1 for i in range(len(xd))], xd, edgecolor='black',linewidth=1, color=color, hatch=hatch, width=0.5)
            plt.xticks([i+1 for i in range(len(y))], y, rotation=40, fontproperties='Times New Roman')
            # plt.yticks([i+1 for i in range(len(x))], x, rotation=40, fontproperties='Times New Roman')
            for xloc, d in enumerate(x):
                xtext = ''.join(map(lambda s:
                                    (','+str(s[1]) if s[0] > 0 else str(s[1])) if (len(str(d))-s[0])%3 == 0 else str(s[1]),
                                    enumerate(str(d))))
                plt.text(xloc+1, int(d/100000)+1,
                         xtext,
                         ha='center',
                         va='bottom',
                         fontdict=font,
                         )
            plt.ylim(ymax=28)
            plt.ticklabel_format(axis="y", style="plain")
            plt.yticks(plt.yticks()[0][1:-1], fontproperties='Times New Roman')
            plt.grid(linestyle='-.', axis='y')
            plt.subplots_adjust(bottom=0.2)
            plt.savefig('%s.pdf' % set)
            plt.show()
            plt.close()



class Exam:

    def Rq1_exam(self):
        # RQ2 在Frequency，average Ochiai下 DNMbfl各倍数下的Exam变化情况
        methods = [
            ['baseline/Sbfl', 'SBFL'],
            ['baseline/Mbfl', 'MBFL'],
            ['baseline/Mcbfl', 'MCBFL'],
            ['HomBaseline/DifferentOperator-1000', 'DifferentOperator'],
            ['HomBaseline/Last2First-1000', 'Last2First'],
            ['HomBaseline/RandomMix-1000', 'RandomMix'],
            ['HomBaseline/Random-0_5', 'Random'],
            ['DNMbfl/DeltaNsMbfl-1', 'DNMBFL'],
        ]
        # fun = 'Tarantula'
        fun = 'Dstar'
        title = ''
        dirpath = './report/CHMBFL/susfile-single/'
        # dirpath = './report/CHMBFL/susfile/'
        data = dict()
        # for file in os.listdir(dirpath):
        for path, name in methods:
            print(name)
            if 'baseline' in path:
                if 'Sbfl' in path:
                    filepath = os.path.join(dirpath, path)
                else:
                    filepath = os.path.join(dirpath, path, 'max')
            else:
                filepath = os.path.join(dirpath, path, 'max')
            exams = []
            top = []
            m = []
            for repeatfile in os.listdir(filepath):
                repeatfilepath = os.path.join(filepath, repeatfile)
                static = readxlsx(repeatfilepath, 'exam-average')
                exams.append(static['form'][fun]['exams'])
                top.append(static['form'][fun]['rank'])
                m.append(static['form'][fun]['map'])

            ave_exam = []
            for i in range(len(exams[0])):
                i_exam = []
                for j in range(len(exams)):
                    i_exam.append(exams[j][i])
                ave_exam.append(sum(i_exam)/len(i_exam))
            data[name] = ave_exam

            ave_top = []
            for i in range(4):
                i_top = []
                for j in range(len(top)):
                    i_top.append(top[j][i])
                ave_top.append(math.ceil(sum(i_top)/len(i_top)))
            print(ave_top, sum(m)/ len(m))

        B = Bline()
        B.date = data
        B.title = title
        B.name = '%s %s %s %s' % ('RQ1', 'Exam', 'Muti', fun)
        # B.exam('./report/CHMBFL/susfile/figure/')
        B.exam(os.path.join(dirpath, 'figure'))

        # box = Box()
        # box.date = data
        # # box.title = '%s-%s-%s-%s' % ('Ochiai', 'Frequency', 'Tie-Average', 'Exam')
        # box.name = '%s-%s-%s-%s-%s' % ('RQ1', 'Box', 'Frequency', 'Average', 'Exam')
        # box.exam('./report/CHMBFL/susfile/figure/')

        return

    def Rq1_exam_new(self):
        # RQ2 在Frequency，average Ochiai下 DNMbfl各倍数下的Exam变化情况
        methods = [
            ['baseline/Sbfl', 'SBFL'],
            ['baseline/Mbfl', 'Metallaxis'],
            ['baseline/Mcbfl', 'MCBFL-hybrid-avg'],
            ['HomBaseline/DifferentOperator-0_5', 'DifferentOperator'],
            ['HomBaseline/Last2First-0_5', 'Last2First'],
            ['HomBaseline/RandomMix-0_5', 'RandomMix'],
            ['HomBaseline/Random-0_5', 'Random'],
            ['DNMbfl/DeltaNsMbfl-1', 'Delta4Ms'],
        ]
        title = ''
        dirpath = './report/CHMBFL/susfile-single/'
        # dirpath = './report/CHMBFL/susfile/'
        for fun in ['Tarantula', 'Ochiai', 'Op2', 'GP13']:
            filename = '%s %s %s %s' % ('RQ1', 'Exam', 'Muti', fun)
            data = dict()
            # for file in os.listdir(dirpath):
            for path, name in methods:
                print(name)
                if 'baseline' in path:
                    if 'Sbfl' in path:
                        filepath = os.path.join(dirpath, path)
                    else:
                        filepath = os.path.join(dirpath, path, 'max')
                else:
                    filepath = os.path.join(dirpath, path, 'max')
                exams = []
                top = []
                m = []
                for repeatfile in os.listdir(filepath):
                    repeatfilepath = os.path.join(filepath, repeatfile)
                    static = readxlsx(repeatfilepath, 'exam-average')
                    exams.append(static['form'][fun]['exams'])
                    top.append(static['form'][fun]['rank'])
                    m.append(static['form'][fun]['map'])

                ave_exam = []
                for i in range(len(exams[0])):
                    i_exam = []
                    for j in range(len(exams)):
                        i_exam.append(exams[j][i])
                    ave_exam.append(sum(i_exam)/len(i_exam))
                data[name] = ave_exam

                ave_top = []
                for i in range(4):
                    i_top = []
                    for j in range(len(top)):
                        i_top.append(top[j][i])
                    ave_top.append(math.ceil(sum(i_top)/len(i_top)))
                print(ave_top, sum(m)/ len(m))

            B = Bline()
            B.date = data
            B.title = title
            B.name = filename
            # B.exam('./report/CHMBFL/susfile/figure/')
            B.exam(os.path.join(dirpath, 'figure'))


        return


    def Rq1_exam_new_point(self):
        # RQ2 在Frequency，average Ochiai下 DNMbfl各倍数下的Exam变化情况
        methods = [
            ['baseline/Sbfl', 'SBFL'],
            ['baseline/Mbfl', 'Metallaxis'],
            ['baseline/Mcbfl', 'MCBFL-hybrid-avg'],
            ['HomBaseline/DifferentOperator-0_5', 'DifferentOperator'],
            # ['HomBaseline/DifferentOperator-1000', 'DifferentOperator'],
            ['HomBaseline/Last2First-0_5', 'Last2First'],
            # ['HomBaseline/Last2First-1000', 'Last2First'],
            ['HomBaseline/RandomMix-0_5', 'RandomMix'],
            # ['HomBaseline/RandomMix-1000', 'RandomMix'],
            ['HomBaseline/Random-0_5', 'Random'],
            ['DNMbfl/DeltaNsMbfl-1', 'Delta4Ms'],
        ]
        title = ''
        dirpath = './report/CHMBFL/susfile-single/'
        # dirpath = './report/CHMBFL/susfile/'
        x = 0.25
        for fun in ['Ochiai']:
            filename = '%s %s %s %s' % ('RQ1', 'Exam', 'Muti', fun)
            data = dict()
            # for file in os.listdir(dirpath):
            for path, name in methods:
                print(name)
                if 'baseline' in path:
                    if 'Sbfl' in path:
                        filepath = os.path.join(dirpath, path)
                    else:
                        filepath = os.path.join(dirpath, path, 'max')
                else:
                    filepath = os.path.join(dirpath, path, 'max')
                exams = []
                top = []
                m = []
                for repeatfile in os.listdir(filepath):
                    repeatfilepath = os.path.join(filepath, repeatfile)
                    static = readxlsx(repeatfilepath, 'exam-average')
                    exams.append(static['form'][fun]['exams'])
                    top.append(static['form'][fun]['rank'])
                    m.append(static['form'][fun]['map'])

                ave_exam = []
                for i in range(len(exams[0])):
                    i_exam = []
                    for j in range(len(exams)):
                        i_exam.append(exams[j][i])
                    ave_exam.append(sum(i_exam)/len(i_exam))
                data[name] = ave_exam

                ave_top = []
                for i in range(4):
                    i_top = []
                    for j in range(len(top)):
                        i_top.append(top[j][i])
                    ave_top.append(math.ceil(sum(i_top)/len(i_top)))
                # print(ave_top, sum(m)/ len(m))

            B = Bline()
            B.date = data
            B.title = title
            B.name = filename
            B.point = x
            # B.exam('./report/CHMBFL/susfile/figure/')
            B.exam(os.path.join(dirpath, 'figure'))


        return

    def Rq3_exam(self):
        # RQ3 分别在Best和Average下 DNMbfl 4种高阶语句怀疑度策略的结果
        # dirpath = './report/CHMBFL/susfile-single'
        dirpath = './report/CHMBFL/susfile/'
        title = ''
        # fun = 'Ochiai'
        fun = 'Tarantula'
        for sheet in ['exam-average']:
            # print(sheet)
            data = dict()
            # for linesus in ['none', 'max', 'ave', 'frequency']:
            for linesus in ['none', 'ave', 'max']:
            # for linesus in ['frequency']:
                print(sheet, linesus)
                top = []
                m = []
                filepath = os.path.join(dirpath, 'DNMbfl', 'DeltaNsMbfl-1.5', linesus)
                exams = []
                for repeatfile in os.listdir(filepath):
                    repeatfilepath = os.path.join(filepath, repeatfile)
                    # print(repeatfilepath)
                    static = readxlsx(repeatfilepath, sheet)
                    exams.append(static['form'][fun]['exams'])
                    top.append(static['form'][fun]['rank'])
                    m.append(static['form'][fun]['map'])
                ave_exam = []
                # for version in exams[0].keys():
                #     for k in range(len(exams[0][version])):
                #         i_exam = []
                #         for j in range(len(exams)):
                #             i_exam.append(exams[j][version][k])
                #         ave_exam.append(sum(i_exam)/len(i_exam))
                for i in range(len(exams[0])):
                    i_exam = []
                    for j in range(len(exams)):
                        i_exam.append(exams[j][i])
                    ave_exam.append(sum(i_exam)/len(i_exam))
                if linesus == 'none':
                    data['frequency'] = ave_exam
                else:
                    data[linesus] = ave_exam

                ave_top = []
                for i in range(4):
                    i_top = []
                    for j in range(len(top)):
                        i_top.append(top[j][i])
                    ave_top.append(math.ceil(sum(i_top)/len(i_top)))
                print(ave_top, sum(m)/ len(m))

            B = Bline()
            B.date = data
            B.title = title
            B.name = '%s-%s-%s-%s' % ('RQ3', 'Bline', 'Average', 'Exam')
            B.exam(os.path.join(dirpath, 'figure'))

            # box = Box()
            # box.date = data
            # # box.title = '%s-%s-%s' % (sheet.split('-')[1], 'Tarantula', 'Exam')
            # box.name = '%s-%s-%s-%s' % ('RQ3', 'Box', 'Average', 'Exam')
            # box.exam('./report/CHMBFL/susfile/figure/')

        return

    def Rq2_exam(self):
        # RQ4 在ave frequency 下比较 （baseline和DNMbfl    挑选部分）在8个公式下的结果
        method = 'DeltaNsMbfl-1'
        dirpath = './report/CHMBFL/susfile-single'
        # dirpath = './report/CHMBFL/susfile/'
        title = ''
        repeat_docfile = os.path.join(dirpath, 'DNMbfl', method, 'max')
        exams = dict()
        top = dict()
        m = dict()
        for repeatfile in os.listdir(repeat_docfile):
            repeatfilepath = os.path.join(repeat_docfile, repeatfile)
            static = readxlsx(repeatfilepath, 'exam-average')
            # for fun in static['form'].keys():
            for fun in [
                'GP13',
                'Ochiai',
                'Op2',
                # 'Dstar',
                # 'Barinel',
                'Tarantula',
            ]:
                # if fun not in ['Tarantula', 'Op2', 'Ochiai', 'Dstar', 'GP13', ]:
                #     continue
                if fun not in exams:
                    exams[fun] = []
                    top[fun] = []
                    m[fun] = []
                exams[fun].append(static['form'][fun]['exams'])
                top[fun].append(static['form'][fun]['rank'])
                m[fun].append(static['form'][fun]['map'])

        ave_exams = dict()
        ave_top = dict()
        for fun in exams.keys():
            ave_exams[fun] = []
            for i in range(len(exams[fun][0])):
                i_exam = []
                for j in range(len(exams[fun])):
                    i_exam.append(exams[fun][j][i])
                ave_exams[fun].append(sum(i_exam)/len(i_exam))

            ave_top[fun] = []
            for i in range(4):
                i_top = []
                for j in range(len(top[fun])):
                    i_top.append(top[fun][j][i])
                ave_top[fun].append(math.ceil(sum(i_top)/len(i_top)))

            print(fun, ave_top[fun], sum(m[fun])/len(m[fun]))

        B = Bline()
        B.date = ave_exams
        B.title = title
        # B.name = '%s-%s-%s-%s-%s' % ('RQ2', 'max', 'Frequency', 'Average', 'Exam')
        B.name = '%s %s %s' % ('RQ2', 'Exam', 'Single')
        B.exam(os.path.join(dirpath, 'figure'))

        # box = Box()
        # box.date = ave_exams
        # # box.title = '%s-%s-%s-%s' % (method, 'frequency', 'average', 'Exam')
        # box.name = '%s-%s-%s-%s-%s' % ('RQ4', 'Box', 'Frequency', 'Average', 'Exam')
        # box.exam('./report/CHMBFL/susfile/figure/')

        # v = Violin()
        # v.date = ave_exams
        # v.title = '%s-%s-%s-%s' % (method, 'frequency', 'average', 'Exam')
        # v.exam('./report/CHMBFL/susfile/figure/')
        return


    def total(self):
        dirpath = './report/CHMBFL/susfile/test'
        readlist = [
            ['./numless/Sbfl/', 'Sbfl', [''], ],
            ['./numless/Mbfl/', 'Mbfl', ['max'], ],
            ['./numless/Mcbfl/', 'Mcbfl', ['max'], ],
            ['./numless/Last2First/', 'L2F', ['max', 'frequency'], ],
            ['./numless/DifferentOperator/', 'DifOp', ['max', 'frequency'], ],
            ['./numless/RandomMix/', 'RMix', ['max', 'frequency'], ],
            ['./numless/Random/', 'Random', ['max', 'frequency'], ],
            # ['./numneed/ErrorDistribution/1', 'ED1', ['max', 'frequency'], ],
            ['./numneed/ErrorDistribution/15', 'ED', ['max', 'frequency'], ],
            ['./numless/MutCluster/', 'MC', ['max', 'frequency'], ],
            # ['./numless/ED+MC/', 'ED+MC', ['max', 'frequency'], ],
        ]

        examorder_list = [['best', 'exam-best'], ['ave', 'exam-average']]
        susfun_list = ['Ochiai', 'Dstar', 'GP13']

        for susfun in susfun_list:
            for examorder, sheet in examorder_list:
                data = dict()
                method_name_list = []
                for path, method_name, homorder_list in readlist:
                    method_name_list.append(method_name)
                    for homorder in homorder_list:
                        if homorder == 'frequency':
                            continue
                        if homorder == '':
                            srcpath = os.path.join(os.getcwd(), './report/CHMBFL/susfile', path)
                        else:
                            srcpath = os.path.join(os.getcwd(), './report/CHMBFL/susfile', path, homorder)
                        for file in os.listdir(srcpath):
                            filepath = os.path.join(srcpath, file)
                            xlsxdata = readxlsx(filepath, sheet)
                            if not xlsxdata:
                                continue
                            if susfun not in xlsxdata:
                                continue
                            # name = '%s\n%s' % (method_name, homorder)
                            name = '%s-%s' % (method_name, homorder)
                            if name not in data:
                                data[name] = []
                            data[name] += xlsxdata[susfun]['exams']


                title = '%s-%s-%s' % (susfun, examorder, 'max')
                photo = Bline()
                photo.title = title
                photo.date = data
                photo.exam(dirpath)
                del data

        return


    def MC(self):
        readlist = [
            ['./numless/MutCluster/', 'MC', ['max', 'frequency'], ],
        ]

        examorder_list = [['best', 'exam-best'], ['ave', 'exam-average']]

        susfun_list = ['Ochiai', 'Dstar', 'GP13']

        for susfun in susfun_list:
            for examorder, sheet in examorder_list:
                data = dict()
                method_name_list = []
                for path, method_name, homorder_list in readlist:
                    method_name_list.append(method_name)
                    for homorder in homorder_list:
                        # if homorder == 'frequency':
                        #     continue
                        if homorder == '':
                            srcpath = os.path.join(os.getcwd(), './report/CHMBFL/susfile', path)
                        else:
                            srcpath = os.path.join(os.getcwd(), './report/CHMBFL/susfile', path, homorder)
                        for file in os.listdir(srcpath):
                            filepath = os.path.join(srcpath, file)
                            xlsxdata = readxlsx(filepath, sheet)
                            if not xlsxdata:
                                continue
                            if susfun not in xlsxdata:
                                continue
                            # name = '%s\n%s' % (method_name, homorder)
                            name = '%s-%s' % (method_name, homorder)
                            if name not in data:
                                data[name] = []
                            data[name] += xlsxdata[susfun]['exams']
                            print(susfun, name, examorder, sum( data[name])/ len( data[name]))


        return



if __name__ == '__main__':

    # Top().total_K1()

    # Top().numneed()
    # Exam().numless()
    # Exam().test()

    # Top().total()
    # Top().total_K1()
    # Exam().total()
    # Exam().MC()

    # Top().total_type()

    # Exam().Rq1_exam()
    # Exam().Rq1_exam_new()
    # Exam().Rq1_exam_new_point()
    # Exam().Rq2_exam()
    # Exam().Rq3_exam()
    # Exam().Rq4_exam()

    Mtp().re3()
    # a = 80.78877960794992
    # b = 65.89923469387755
    # print(a-b)

