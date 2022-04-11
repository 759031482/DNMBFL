from scipy.stats import ttest_ind
from scipy.stats import ks_2samp
from scipy.stats import wilcoxon
from scipy.stats import levene
from scipy.stats import kstest
import os
import openpyxl

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import scipy.stats as st
import seaborn as sns

from codeflaws_version_control import readxlsx
from codeflaws_version_control import Qr_excel


def readxlsx_e(path, sheet):
    try:
        data_read = Qr_excel().read(path, sheet)
    except:
        print('读取失败 %s %s' % (path, sheet))
        return False
    static_all = {
        'totalfaultnum': 0,
        'ttime': 0,
        'fomnum': 0,
        'homnum': 0,
        'Operator': [0 for _ in range(13)],
        'Accurate': [0 for _ in range(3)],
        'versionnum': len(data_read)-1,
        'form': dict()
    }
    for i in range(len(data_read)):
        row_data = data_read[i]
        for key, value in row_data['self'].items():
            if value == '':
                continue
            if key == 'Fault_Record':
                static_all['totalfaultnum'] += int(value)
            elif key == 'TimeSpend':
                static_all['ttime'] += int(value)
            elif key == 'Fomnum':
                try:
                    static_all['fomnum'] += int(value)
                except:
                    print(row_data['self'])
            elif key == 'Homnum':
                static_all['homnum'] += int(value)
            elif 'Operator' in key:
                OperatorType = int(key.split(' ')[2])
                static_all['Operator'][OperatorType] = static_all['Operator'][OperatorType] + value*100
            elif 'Accurate' in key:
                if key == 'None Accurate':
                    static_all['Accurate'][0] += value*100
                if key == 'Pare Accurate':
                    static_all['Accurate'][1] += value*100
                if key == 'Accurate':
                    static_all['Accurate'][2] += value*100
        for key, value in row_data['form'].items():
            if key not in static_all['form']:
                static_all['form'][key] = {
                    'exam': 0,
                    # 'exams': dict(),
                    'exams': [],
                    'rank': [0 for _ in range(4)],
                    'map': 0,
                }
            exam = 0
            for j in range(len(row_data['form'][key]['exam'])):
                exam += row_data['form'][key]['exam'][j]
                rank = row_data['form'][key]['rank'][j]
                # static_all['form'][key]['exams'].append(exam)
                static_all['form'][key]['map'] += 1/rank
                if rank <= 1:
                    static_all['form'][key]['rank'][0] += 1
                if rank <= 3:
                    static_all['form'][key]['rank'][1] += 1
                if rank <= 5:
                    static_all['form'][key]['rank'][2] += 1
                if rank <= 10:
                    static_all['form'][key]['rank'][3] += 1
            static_all['form'][key]['exam'] += exam/len(row_data['form'][key]['exam'])
            static_all['form'][key]['exams'].append(exam/len(row_data['form'][key]['exam']))

    static_all['fomnum'] = static_all['fomnum']/static_all['versionnum']
    static_all['homnum'] = static_all['homnum']/static_all['versionnum']
    for i in range(13):
        static_all['Operator'][i] = static_all['Operator'][i]/static_all['versionnum']
    for i in range(3):
        static_all['Accurate'][i] = static_all['Accurate'][i]/static_all['versionnum']
    for fun in static_all['form'].keys():
        if static_all['totalfaultnum'] == 0:
            static_all['form'][fun]['exam'] = 0
            static_all['form'][fun]['map'] = 0
        else:
            static_all['form'][fun]['exam'] = static_all['form'][fun]['exam']/static_all['totalfaultnum']
            static_all['form'][fun]['map'] = static_all['form'][fun]['map']/static_all['versionnum']

    return static_all




def rq1():

    wb = openpyxl.Workbook()
    del wb['Sheet']

    sustype = 'max'
    tietype = 'ave'

    doc_path = './report/CHMBFL/susfile-single'
    save_path = './HypothesisReport/rq1-single.xlsx'
    # doc_path = './report/CHMBFL/susfile'
    # save_path = './HypothesisReport/rq1-muti.xlsx'
    methodlist = [
        ['Sbfl', 'baselines', 'baseline/Sbfl'],
        ['Mbfl', 'baselines', 'baseline/Mbfl'],
        ['Mcbfl', 'baselines', 'baseline/Mcbfl'],
        ['Last2First', 'baselines', 'HomBaseline/Last2First-0_5'],
        ['DifferentOperator', 'baselines', 'HomBaseline/DifferentOperator-0_5'],
        ['RandomMix', 'baselines', 'HomBaseline/RandomMix-0_5'],
        ['Random', 'baselines', 'HomBaseline/Random-0_5'],
        ['DeltaNsMbfl', '', 'DNMbfl/DeltaNsMbfl-1'],
    ]

    data = dict()
    avedata = dict()
    funlist = ['GP13', 'Ochiai', 'Op2', 'Tarantula']
    for fun in funlist:
        data[fun] = dict()
        avedata[fun] = dict()
        for baseline, describe, baseline_path in methodlist:
            data[fun][baseline] = []
            avedata[fun][baseline] = []

    namelist = []
    for baseline, describe, baseline_path in methodlist:
        namelist.append(baseline)
        print(baseline)
        path = os.path.join(os.getcwd(), doc_path, baseline_path, sustype)
        if baseline == 'Sbfl':
            path = os.path.join(os.getcwd(), doc_path, baseline_path)
        elif baseline == 'Mbfl' or baseline == 'Mcbfl':
            path = os.path.join(os.getcwd(), doc_path, baseline_path, 'max')

        if not os.path.exists(path):
            print('路径不存在 %s' % path)
            continue

        for repeatfile in os.listdir(path):
            repeatfilepath = os.path.join(path, repeatfile)
            static = readxlsx(repeatfilepath, 'exam-average')
            for fun in funlist:
                data[fun][baseline].append(static['form'][fun]['exams'])

    print('数据获取完成')

    for fun in funlist:
        for baseline, describe, baseline_path in methodlist:
            data_fun_baseline = data[fun][baseline]
            ave = []
            for i in range(len(data_fun_baseline[0])):
                i_exam = []
                for j in range(len(data_fun_baseline)):
                    i_exam.append(data_fun_baseline[j][i])
                ave.append(sum(i_exam)/len(i_exam))
            avedata[fun][baseline] = ave

    for fun in funlist:
        inspectionable = True
        for baseline1 in avedata[fun].keys():
            for baseline2 in avedata[fun].keys():
                if not len(avedata[fun][baseline1]) == len(avedata[fun][baseline2]):
                    inspectionable = False
                    print(baseline1, baseline2, '长度不等')
        if not inspectionable:
            return False
        ws = wb.create_sheet(fun)
        Plist = []
        for baseline1 in avedata[fun].keys():
            Plist.append([])
            for baseline2 in avedata[fun].keys():
                if baseline1 == baseline2:
                    Plist[-1].append(1)
                    continue
                res = wilcoxon(avedata[fun][baseline1], avedata[fun][baseline2], alternative='less').pvalue
                # Plist[-1].append(float('%.3f'%res))
                Plist[-1].append(res)
        for i in range(len(Plist)):
            print(namelist[i], Plist[i])

        for i in range(len(namelist)):
            ws.cell(i+2, 1, namelist[i])
            ws.cell(1, i+2, namelist[i])
        for i in range(len(namelist)):
            for j in range(len(namelist)):
                if i == j:
                    ws.cell(i+2, j+2, '\\')
                else:
                    ws.cell(i+2, j+2, Plist[i][j])
    wb.save(save_path)
    print(save_path)

    return Plist, namelist


def rq1_new():

    wb = openpyxl.Workbook()

    for fun in ['Tarantula']:
        ws = wb.create_sheet(fun)
        sustype = 'max'
        tietype = 'ave'

        doc_path = './report/CHMBFL/susfile'
        # doc_path = './report/CHMBFL/susfile-single'
        # save_path = './HypothesisReport/rq1-single.xlsx'
        save_path = './HypothesisReport/rq3-muti-h.xlsx'
        methodlist = [
            ['DNMBFL-0', '', 'DNMbfl/DeltaNsMbfl-0'],
            ['DNMBFL-0.5', '', 'DNMbfl/DeltaNsMbfl-0.5'],
            ['DNMBFL-1', '', 'DNMbfl/DeltaNsMbfl-1'],
            ['DNMBFL-1.5', '', 'DNMbfl/DeltaNsMbfl-1.5'],
            ['DNMBFL-2', '', 'DNMbfl/DeltaNsMbfl-2'],
            ['DNMBFL-2.5', '', 'DNMbfl/DeltaNsMbfl-2.5'],
            ['DNMBFL-3', '', 'DNMbfl/DeltaNsMbfl-3'],
            ['DNMBFL-3.5', '', 'DNMbfl/DeltaNsMbfl-3.5'],
            ['DNMBFL-4', '', 'DNMbfl/DeltaNsMbfl-4']
        ]

        namelist = []
        data = dict()
        for baseline, describe, baseline_path in methodlist:
            namelist.append(baseline)
            print(baseline)
            path = os.path.join(os.getcwd(), doc_path, baseline_path, sustype)
            if not sustype == 'max' and (baseline == 'Mbfl' or baseline == 'Mcbfl' or baseline == 'DeltaMbfl'):
                path = os.path.join(os.getcwd(), doc_path, baseline_path, 'max')
            if baseline == 'Sbfl':
                path = os.path.join(os.getcwd(), doc_path, baseline_path)

            if not os.path.exists(path):
                print('路径不存在 %s' % path)
                continue

            exams = []
            for repeatfile in os.listdir(path):
                repeatfilepath = os.path.join(path, repeatfile)
                static = readxlsx(repeatfilepath, 'exam-average')
                exams.append(static['form'][fun]['exams'])
                # break

            ave_exam = []
            for i in range(len(exams[0])):
                i_exam = []
                for j in range(len(exams)):
                    i_exam.append(exams[j][i])
                ave_exam.append(sum(i_exam)/len(i_exam))
            data[baseline] = ave_exam
        print('数据获取完成')

        inspectionable = True
        for baseline1 in data.keys():
            for baseline2 in data.keys():
                if not len(data[baseline1]) == len(data[baseline2]):
                    inspectionable = False
                    print(baseline1, baseline2, '长度不等')
        if not inspectionable:
            return False


        fig,ax = plt.subplots(nrows=3,ncols=3,figsize=(10,6))
        for i, baseline1 in enumerate(data.keys()):
            ave = sum(data[baseline1])/len(data[baseline1])
            dev = sum(map(lambda x: (x-ave)*(x-ave), data[baseline1]))
            sns.kdeplot(x=data[baseline1], ax=ax[i//3][i%3])
            print(baseline1, ave, dev, kstest(data[baseline1], 'norm'))
        plt.show()

        Plist = []
        for baseline1 in data.keys():
            Plist.append([])
            for baseline2 in data.keys():
                if baseline1 == baseline2:
                    Plist[-1].append(1)
                    continue
                # res = wilcoxon(data[baseline1], data[baseline2]).pvalue
                res = levene(data[baseline1], data[baseline2]).pvalue
                # Plist[-1].append(float('%.3f'%res))
                Plist[-1].append(res)
        for i in range(len(Plist)):
            print(namelist[i], Plist[i])

        for i in range(len(namelist)):
            ws.cell(i+2, 1, namelist[i])
            ws.cell(1, i+2, namelist[i])
        for i in range(len(namelist)):
            for j in range(len(namelist)):
                if i == j:
                    ws.cell(i+2, j+2, '\\')
                else:
                    ws.cell(i+2, j+2, Plist[i][j])
    # wb.save(save_path)

    return Plist, namelist


def rq1_new2():

    wb = openpyxl.Workbook()

    for fun in ['Tarantula']:
        ws = wb.create_sheet(fun)
        sustype = 'max'
        tietype = 'ave'

        doc_path = './report/CHMBFL/susfile'
        # doc_path = './report/CHMBFL/susfile-single'
        # save_path = './HypothesisReport/rq1-single.xlsx'
        save_path = './HypothesisReport/rq3-muti-h.xlsx'
        methodlist = [
            ['DNMBFL-0', '', 'DNMbfl/DeltaNsMbfl-0'],
            ['DNMBFL-0.5', '', 'DNMbfl/DeltaNsMbfl-0.5'],
            ['DNMBFL-1', '', 'DNMbfl/DeltaNsMbfl-1'],
            ['DNMBFL-1.5', '', 'DNMbfl/DeltaNsMbfl-1.5'],
            ['DNMBFL-2', '', 'DNMbfl/DeltaNsMbfl-2'],
            ['DNMBFL-2.5', '', 'DNMbfl/DeltaNsMbfl-2.5'],
            ['DNMBFL-3', '', 'DNMbfl/DeltaNsMbfl-3'],
            ['DNMBFL-3.5', '', 'DNMbfl/DeltaNsMbfl-3.5'],
            ['DNMBFL-4', '', 'DNMbfl/DeltaNsMbfl-4']
        ]

        namelist = []
        data = dict()
        data_ave = dict()
        data_dev = dict()
        for baseline, describe, baseline_path in methodlist:
            namelist.append(baseline)
            print(baseline)
            path = os.path.join(os.getcwd(), doc_path, baseline_path, sustype)
            if not sustype == 'max' and (baseline == 'Mbfl' or baseline == 'Mcbfl' or baseline == 'DeltaMbfl'):
                path = os.path.join(os.getcwd(), doc_path, baseline_path, 'max')
            if baseline == 'Sbfl':
                path = os.path.join(os.getcwd(), doc_path, baseline_path)

            if not os.path.exists(path):
                print('路径不存在 %s' % path)
                continue

            exams = []
            for repeatfile in os.listdir(path):
                repeatfilepath = os.path.join(path, repeatfile)
                static = readxlsx(repeatfilepath, 'exam-average')
                exams.append(static['form'][fun]['exams'])
                # break

            dev_list = []
            ave_list = []
            for j in range(len(exams[0])):
                for i in range(len(exams)):
                    ave = sum(map(lambda x: x[j], exams))/len(exams)
                    dev = sum(map(lambda x: abs(x[j]-ave)/ave, exams))/len(exams)
                    dev_list.append(dev)
                    ave_list.append(ave)

            # ave_exam = []
            # for i in range(len(exams[0])):
            #     i_exam = []
            #     for j in range(len(exams)):
            #         i_exam.append(exams[j][i])
            #     ave_exam.append(sum(i_exam)/len(i_exam))
            # data[baseline] = ave_exam
            data_ave[baseline] = sum(ave_list)/len(ave_list)
            data_dev[baseline] = sum(dev_list)/len(dev_list)
        print('数据获取完成')
        for baseline in data_dev.keys():
            # print(baseline, data_ave[baseline], data_dev[baseline])
            print(baseline, data_ave[baseline], data_dev[baseline])

        inspectionable = True
        for baseline1 in data.keys():
            for baseline2 in data.keys():
                if not len(data[baseline1]) == len(data[baseline2]):
                    inspectionable = False
                    print(baseline1, baseline2, '长度不等')
        # if not inspectionable:
        #     return False


        # fig,ax = plt.subplots(nrows=3,ncols=3,figsize=(10,6))
        # for i, baseline1 in enumerate(data.keys()):
        #     ave = sum(data[baseline1])/len(data[baseline1])
        #     dev = sum(map(lambda x: (x-ave)*(x-ave), data[baseline1]))
        #     sns.kdeplot(x=data[baseline1], ax=ax[i//3][i%3])
        #     print(baseline1, ave, dev, kstest(data[baseline1], 'norm'))
        # plt.show()

        # Plist = []
        # for baseline1 in data.keys():
        #     Plist.append([])
        #     for baseline2 in data.keys():
        #         if baseline1 == baseline2:
        #             Plist[-1].append(1)
        #             continue
        #         # res = wilcoxon(data[baseline1], data[baseline2]).pvalue
        #         res = levene(data[baseline1], data[baseline2]).pvalue
        #         # Plist[-1].append(float('%.3f'%res))
        #         Plist[-1].append(res)
        # for i in range(len(Plist)):
        #     print(namelist[i], Plist[i])
        #
        # for i in range(len(namelist)):
        #     ws.cell(i+2, 1, namelist[i])
        #     ws.cell(1, i+2, namelist[i])
        # for i in range(len(namelist)):
        #     for j in range(len(namelist)):
        #         if i == j:
        #             ws.cell(i+2, j+2, '\\')
        #         else:
        #             ws.cell(i+2, j+2, Plist[i][j])
    # wb.save(save_path)

    return
    # return Plist, namelist


def rq2():
    sustype = 'max'
    tietype = 'ave'

    funlist = ['GP13', 'Ochiai', 'Op2', 'Tarantula']

    # doc_path = './report/CHMBFL/susfile'
    # savepath = './HypothesisReport/rq2-muti.xlsx'
    doc_path = './report/CHMBFL/susfile-single'
    savepath = './HypothesisReport/rq2-single.xlsx'

    baseline = 'DeltaNsMbfl'
    baseline_path = 'DNMbfl/DeltaNsMbfl-1'

    namelist = []
    data = dict()
    avedata = dict()
    for fun in funlist:
        data[fun] = []
        avedata[fun] = []
        namelist.append(fun)

    path = os.path.join(os.getcwd(), doc_path, baseline_path, sustype)
    if baseline == 'Sbfl':
        path = os.path.join(os.getcwd(), doc_path, baseline_path)
    elif baseline == 'Mbfl' or baseline == 'Mcbfl':
        path = os.path.join(os.getcwd(), doc_path, baseline_path, 'max')

    if not os.path.exists(path):
        print('路径不存在 %s' % path)
        return

    for repeatfile in os.listdir(path):
        repeatfilepath = os.path.join(path, repeatfile)
        static = readxlsx(repeatfilepath, 'exam-average')
        for fun in funlist:
            data[fun].append(static['form'][fun]['maps'])

    for fun in funlist:
        data_fun = data[fun]
        ave = []
        for i in range(len(data_fun[0])):
            i_exam = []
            for j in range(len(data_fun)):
                i_exam.append(data_fun[j][i])
            ave.append(sum(i_exam)/len(i_exam))
        avedata[fun] = ave
    print('数据获取完成')

    inspectionable = True
    for key1 in avedata.keys():
        for key2 in avedata.keys():
            if not len(avedata[key1]) == len(avedata[key2]):
                inspectionable = False
                print(key1, key2, '长度不等')
    if not inspectionable:
        return False

    Plist = []
    for baseline1 in avedata.keys():
        Plist.append([])
        for baseline2 in avedata.keys():
            if baseline1 == baseline2:
                Plist[-1].append(1)
                continue
            res = wilcoxon(avedata[baseline1], avedata[baseline2], alternative='greater').pvalue
            Plist[-1].append(float('%.3f'%res))
    for i in range(len(Plist)):
        print(namelist[i], Plist[i])

    wb = openpyxl.Workbook()
    ws = wb['Sheet']
    for i in range(len(namelist)):
        ws.cell(i+2, 1, namelist[i])
        ws.cell(1, i+2, namelist[i])
    for i in range(len(namelist)):
        for j in range(len(namelist)):
            if i == j:
                ws.cell(i+2, j+2, '\\')
            else:
                ws.cell(i+2, j+2, Plist[i][j])
    wb.save(savepath)
    print(savepath)

    return


def rq3():
    tietype = 'ave'
    fun = 'Tarantula'
    sustypelist = ['max', 'ave', 'none']

    doc_path = './report/CHMBFL/susfile-single'
    savepath = './HypothesisReport/rq3-single.xlsx'
    baseline = 'DeltaNsMbfl'
    baseline_path = 'DNMbfl/DeltaNsMbfl-1.5'

    namelist = []
    data = dict()
    for sustype in sustypelist:
        namelist.append(sustype)
        print(sustype)
        path = os.path.join(os.getcwd(), doc_path, baseline_path, sustype)
        if not sustype == 'max' and (baseline == 'Mbfl' or baseline == 'Mcbfl' or baseline == 'DeltaMbfl'):
            path = os.path.join(os.getcwd(), doc_path, baseline_path, 'max')
        if baseline == 'Sbfl':
            path = os.path.join(os.getcwd(), doc_path, baseline_path)

        if not os.path.exists(path):
            print('路径不存在 %s' % path)
            continue

        exams = []
        for repeatfile in os.listdir(path):
            repeatfilepath = os.path.join(path, repeatfile)
            static = readxlsx(repeatfilepath, 'exam-average')
            exams.append(static['form'][fun]['exams'])

        ave_exam = []
        for i in range(len(exams[0])):
            i_exam = []
            for j in range(len(exams)):
                i_exam.append(exams[j][i])
            ave_exam.append(sum(i_exam)/len(i_exam))
        data[sustype] = ave_exam
    print('数据获取完成')

    inspectionable = True
    for key1 in data.keys():
        for key2 in data.keys():
            if not len(data[key1]) == len(data[key2]):
                inspectionable = False
                print(key1, key2, '长度不等')
    if not inspectionable:
        return False

    Plist = []
    for baseline1 in data.keys():
        Plist.append([])
        for baseline2 in data.keys():
            if baseline1 == baseline2:
                Plist[-1].append(1)
                continue
            res = wilcoxon(data[baseline1], data[baseline2]).pvalue
            Plist[-1].append(float('%.4f'%res))
    for i in range(len(Plist)):
        print(namelist[i], Plist[i])

    wb = openpyxl.Workbook()
    ws = wb['Sheet']
    for i in range(len(namelist)):
        ws.cell(i+2, 1, namelist[i])
        ws.cell(1, i+2, namelist[i])
    for i in range(len(namelist)):
        for j in range(len(namelist)):
            if i >= j:
                ws.cell(i+2, j+2, '\\')
            else:
                ws.cell(i+2, j+2, Plist[i][j])
    wb.save(savepath)

    return


def rq4():
    strategylist = [
        ['Sbfl', '', 'Tarantula'],
        ['Mbfl', 'max', 'Tarantula'],
        ['Mcbfl', 'max', 'Tarantula'],
        ['Last2First', '', 'Tarantula'],
        ['DifferentOperator', '', 'Tarantula'],
        ['RandomMix', '', 'Tarantula'],
        ['Random', '', 'Tarantula'],
        ['DeltaNsMbfl', '', 'Tarantula'],
    ]
    tietype = 'ave'
    fun = 'Tarantula'
    sustypelist = ['max', 'ave', 'none']

    doc_path = './report/CHMBFL/susfile'
    baseline = 'DeltaNsMbfl'
    baseline_path = 'DNMbfl/DeltaNsMbfl-1.5'

    namelist = []
    data = dict()
    for sustype in sustypelist:
        namelist.append(sustype)
        print(sustype)
        path = os.path.join(os.getcwd(), doc_path, baseline_path, sustype)
        if not sustype == 'max' and (baseline == 'Mbfl' or baseline == 'Mcbfl' or baseline == 'DeltaMbfl'):
            path = os.path.join(os.getcwd(), doc_path, baseline_path, 'max')
        if baseline == 'Sbfl':
            path = os.path.join(os.getcwd(), doc_path, baseline_path)

        if not os.path.exists(path):
            print('路径不存在 %s' % path)
            continue

        exams = []
        for repeatfile in os.listdir(path):
            repeatfilepath = os.path.join(path, repeatfile)
            static = readxlsx(repeatfilepath, 'exam-average')
            exams.append(static['form'][fun]['exams'])

        ave_exam = []
        for i in range(len(exams[0])):
            i_exam = []
            for j in range(len(exams)):
                i_exam.append(exams[j][i])
            ave_exam.append(sum(i_exam)/len(i_exam))
        data[sustype] = ave_exam
    print('数据获取完成')

    inspectionable = True
    for key1 in data.keys():
        for key2 in data.keys():
            if not len(data[key1]) == len(data[key2]):
                inspectionable = False
                print(key1, key2, '长度不等')
    if not inspectionable:
        return False

    Plist = []
    for baseline1 in data.keys():
        Plist.append([])
        for baseline2 in data.keys():
            if baseline1 == baseline2:
                Plist[-1].append(1)
                continue
            res = wilcoxon(data[baseline1], data[baseline2]).pvalue
            Plist[-1].append(float('%.4f'%res))
    for i in range(len(Plist)):
        print(namelist[i], Plist[i])

    wb = openpyxl.Workbook()
    ws = wb['Sheet']
    for i in range(len(namelist)):
        ws.cell(i+2, 1, namelist[i])
        ws.cell(1, i+2, namelist[i])
    for i in range(len(namelist)):
        for j in range(len(namelist)):
            if i >= j:
                ws.cell(i+2, j+2, '\\')
            else:
                ws.cell(i+2, j+2, Plist[i][j])
    wb.save('./HypothesisReport/rq3.xlsx')

    return


def test():
    x1 = [1for i in range(100)]
    x2 = [2for i in range(100)]

    res = wilcoxon(x1, x2, alternative='greater').pvalue

    print(res)

if __name__ == '__main__':
    rq1()
    # rq2()
    # rq3()
    # rq1_new()
    # rq1_new2()
    # test()
